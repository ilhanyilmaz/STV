from openpyxl import Workbook
from openpyxl import load_workbook
import random
import sys
import math

wbFilename = "deneme.xlsx"
numAday = 6
numSandalye = 2
numBireyOy = 2
numSecmen = 20
mQuota = 0
lineStr=""
electedCandidates = []
eliminatedCandidates = []

def createOylar():
	#wbFilename = raw_input("Dosya adini gir: ")
	#numAday = int(raw_input("Aday sayisini gir: "))
	#numSandalye = int(raw_input("Sandalye sayisini gir: "))
	#numBireyOy = int(raw_input("Bir kisinin verebilecegi maksimum oy sayisini gir: "))
	#numSecmen = int(raw_input("Gecerli pusula sayisini gir: "))

	global numAday
	global numSecmen
	global numBireyOy
	print("Oylar uretiliyor")

	tumOylar = {}

	for i in range(numSecmen):
		numSecmenOy = (random.randrange(1,numBireyOy+1))
		adayOylar = []
		sAdayOylar = ""
		for j in range(numSecmenOy):
			adayNo = random.randrange(numAday)
			while adayNo in adayOylar:
				adayNo = random.randrange(numAday)
			adayOylar.append(adayNo)
			sAdayOylar = sAdayOylar + "," + str(adayNo)
		
		sAdayOylar = sAdayOylar[1:]

		if not sAdayOylar in tumOylar.keys() : 
			tumOylar[sAdayOylar] = 1
		else :
			tumOylar[sAdayOylar] += 1
		
	print tumOylar
	return tumOylar

def getOy(keyStr, oyNo):
	oylar = keyStr.split(',')
	return int(oylar[oyNo])

def removeOy(keyStr):
	oylar = keyStr.split(',')
	newKeyStr = ""
	for i in range(len(oylar) - 1):
		newKeyStr = newKeyStr + "," + oylar[i+1]
	newKeyStr = newKeyStr[1:]
	return newKeyStr

def adayOylar(tumOylar):
	global numAday
	global lineStr
	firstRoundOylar = []
	for i in range(numAday):
		firstRoundOylar.append(0)
	oylarKeys = tumOylar.keys()
	for i in range(len(oylarKeys)):
		secilen = getOy(oylarKeys[i], 0)
		firstRoundOylar[secilen] += tumOylar[oylarKeys[i]]

	lineStr = str(firstRoundOylar)
	print firstRoundOylar
	return firstRoundOylar

def lowestIndex(oyToplam):
	global numSecmen
	nLowestIndex = -1
	lowest = numSecmen
	for i in range(len(oyToplam)) :
		if oyToplam[i] > 0 and oyToplam[i] < lowest :
			nLowestIndex = i
			lowest = oyToplam[i]
	print "En dusuk oy sahibi " + str(nLowestIndex) + " nolu aday."
	print "Birden fazla en dusuk varsa napilacak? Gecerli pusula sayisi dusurulecek mi?"
	return nLowestIndex

def removeAday(tumOylar, adayNo):
	oylarKeys = tumOylar.keys()
	for i in range(len(oylarKeys)):
		if getOy(oylarKeys[i],0) == adayNo:
			tempKeys = oylarKeys[i]
			newKeyStr = removeOy(tempKeys)
			if newKeyStr in oylarKeys:
				tumOylar[newKeyStr] += tumOylar[tempKeys]
			elif not newKeyStr == "" :
				tumOylar[newKeyStr] = tumOylar[tempKeys]

			del tumOylar[tempKeys]
	return

def sumOylar(roundOylar):
	nSumOylar = 0
	for i in range(len(roundOylar)):
		nSumOylar += roundOylar[i]
	print nSumOylar
	return nSumOylar

def leftEnoughAday(oyToplam):
	global numSandalye
	numRemAday = 0
	for i in range(len(oyToplam)) :
		if oyToplam[i] > 0 :
			numRemAday+= 1

	if numRemAday > numSandalye:
		return True
	else:
		return False

def checkRound(tumOylar):
	global mQuota

	oyToplam = adayOylar(tumOylar)

	if not leftEnoughAday(oyToplam) :
		return True

	hasSurplus = False
	surplusList = []
	for i in range(len(oyToplam)):
		if oyToplam[i] > mQuota :
			hasSurplus = True
			surplusList.append(i)

	if not hasSurplus :
		nLowestIndex = lowestIndex(oyToplam)
		removeAday(tumOylar, nLowestIndex)
		print "Removing Aday: Add votes to 2nd choices"
		return False
	else :
		for i in range(len(surplusList)):
			surplusIndex = surplusList[i]
			ratio = oyToplam[surplusIndex] / mQuota
			print "There is surplus, write the code god dammit"
		return True



def loadSecimDegiskenleri(wb):
	global numSandalye
	global numAday
	global numSecmen
	global numBireyOy
	wsAyarlar = wb['Ayarlar']
	numSandalye = int(wsAyarlar['B1'].value)
	numAday = int(wsAyarlar['B2'].value)
	numSecmen = int(wsAyarlar['B3'].value)
	numBireyOy = int(wsAyarlar['B4'].value)
	return


def saveSecimDegiskenleri(wb):
	global numSandalye
	global numAday
	global numSecmen
	global numBireyOy
	wsAyarlar = wb.active
	wsAyarlar.title = 'Ayarlar'
	wsAyarlar['A1'] = "Sandalye sayisi"
	wsAyarlar['B1'] = numSandalye
	wsAyarlar['A2'] = "Aday sayisi"
	wsAyarlar['B2'] = numAday
	wsAyarlar['A3'] = "Gecerli pusula sayisi"
	wsAyarlar['B3'] = numSecmen
	wsAyarlar['A4'] = "Bireysel tercih sayisi"
	wsAyarlar['B4'] = numBireyOy
	return

def saveOylar(wb, tumOylar):
	wsOylar = wb.create_sheet(title="Oylar")
	oylarKeys = tumOylar.keys()
	for i in range(len(oylarKeys)):
		cellName = "A"+str(i+1)
		keyName = oylarKeys[i]
		wsOylar[cellName] = keyName
		cellName = "B"+str(i+1)
		wsOylar[cellName] = tumOylar[keyName]
	return

def loadOylar(wb):
	wsOylar = wb['Oylar']
	tumOylar = {}
	i=0
	while True:
		i+=1
		key = str(wsOylar["A"+str(i)].value)
		countStr = wsOylar["B"+str(i)].value
		if countStr == None:
			break
		tumOylar[key]= int(countStr)
	return tumOylar




def hasOngoingCandidate(ballotPaper):
	global electedCandidates
	global eliminatedCandidates
	for i in range(len(ballotPaper)):
		if ballotPaper[i] not in electedCandidates and ballotPaper[i] not in eliminatedCandidates :
			return True
	return False

def getContinuingCandidate(ballotPaper):
	global electedCandidates
	global eliminatedCandidates
	for i in range(len(ballotPaper)):
		if ballotPaper[i] not in electedCandidates and ballotPaper[i] not in eliminatedCandidates :
			return ballotPaper[i]
	return -1

def newArrayWithZeros(size):
	array = []
	for i in size:
		array.append(0)


def main(argv):
	global mQuota
	global numSandalye
	global numSecmen
	global lineStr
	global numAday
	
	ballotPaperValues = []
	candidateVotes = newArrayWithZeros(numAday)
	
	if(len(argv)>1):
		wb = load_workbook(filename = argv[1])
		loadSecimDegiskenleri(wb)
		tumOylar = loadOylar(wb)
	else:
		wb = Workbook()
		saveSecimDegiskenleri(wb)
		tumOylar = createOylar()
		saveOylar(wb, tumOylar)

	for i in range(len(tumOylar)):
		if hasOngoingCandidate(tumOylar[i]) == False:
			#(a) Any ballot paper that does not express a transferable preference for a continuing candidate is declared exhausted-without-value and recorded with a value of zero.
			ballotPaperValues[i] = 0
		else :
			#(b) Assign each ballot paper allocated to a Continuing Candidate a Vote value of one.
			candidate = getContinuingCandidate(tumOylar[i])
			candidateVotes[candidate] += 1


	#(c) Ascertain and assign the Candidate’s Total Value of the Vote (Ctvv) for each Continuing Candidate by aggregating the value of the votes allocated to each Continuing Candidate.
	cctvs = adayOylar(tumOylar) # Candidate's total value of votes
	#(d) Ascertain and assign the Total Vote (Tv) value by aggregating the total value of votes allocated to each candidate outlined in (c) above.
	tv = sumOylar(cctvs)
	#(e) Calculate the Quota required to elect a candidate by dividing the Total Vote (Tv) by 1 more than the number of candidates required to be elected and by increasing the quotient (disregarding any remainder) by 1 (Q = integer(Tv/(1+No of Vacancies))+1).
	mQuota = int(math.floor((tv/ (numSandalye-len(electedCandidates)+1) + 1)))


	#Provisional declaration of elected candidates
	#Any candidate who has received a Total value of votes (Ctvv) equal to or greater than the Quota (Q) is to be provisionally declared elected.

	for i in range(len(cctvs)):
		if cctvs[i]>=mQuota:
			electedCandidates.append(i)

	#######https://en.wikipedia.org/wiki/Wright_system#Definitions



	wsRounds = wb.create_sheet("Rounds")


	mQuota = int(math.floor((numSecmen/ (numSandalye+1) + 1)))
	print "Kota: " + str(mQuota)

	roundNo = 1
	while not checkRound(tumOylar):
		wsRounds["A"+str(roundNo)] = lineStr
		print "Next round"
		roundNo += 1
	wb.save(filename = str(wbFilename))

	print argv
	return

if __name__ == "__main__":
	main(sys.argv)